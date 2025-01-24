from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
import time 
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import workbook, load_workbook
import pandas as pd 
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime

planilha = load_workbook('A RECEBER 1.xlsx')
plan = planilha.active
plan = planilha.worksheets[0]

planilha = load_workbook('TRANSFERENCIA 1.xlsx')
plan = planilha.active
plan = planilha.worksheets[0]

chrome_options = Options()
chrome_options.add_argument("--headless")   

def login(navejgar, user, password):
    try:
        WebDriverWait(navejgar, 5).until(EC.element_to_be_clickable((By.ID, 'Username')))
        inpulogin = navejgar.find_element(By.ID, 'Username')
        inpulogin.clear()
        inpulogin.send_keys(user)

        inputpass = navejgar.find_element(By.ID, "Password")
        inputpass.clear()
        inputpass.send_keys(password)
        time.sleep(5)

        botton = navejgar.find_element(By.ID, 'SignIn')
        botton.click()
        time.sleep(2)
    except:
        print("Erro ao fazer o login")

def preencher_campos_areceber(navejgar, plan):
    try:
        # Preencher proprietário
        prop = WebDriverWait(navejgar, 10).until(EC.element_to_be_clickable((By.ID, 'ProprietarioText')))
        prop.clear()
        prop.send_keys(plan['A2'].value + Keys.ENTER)
       
        # Preencher Devedor
        devedor = WebDriverWait(navejgar, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="DevedorText"]')))
        devedor.clear()
        navejgar.execute_script("arguments[0].scrollIntoView(true);", devedor)
        devedor.send_keys(plan['B2'].value + Keys.ENTER)
        time.sleep(1)

        # Preencher descrição
        descricao = WebDriverWait(navejgar, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="Descricao"]')))
        navejgar.execute_script("arguments[0].scrollIntoView(true);", descricao)
        descricao.clear()
        descricao.send_keys(plan['C2'].value)
        time.sleep(5)

        # Formatar a data no formato desejado
        if plan['D2'].value is not None:
            data_emissao = plan['D2'].value
        # Converter a data para o formato desejado
            if isinstance(data_emissao, datetime):
                data_formatada = data_emissao.strftime('%d/%m/%Y')
            else:
                print(f"Formato de data inesperado: {data_emissao}")
                return
            
            dateEm = WebDriverWait(navejgar, 20).until(EC.element_to_be_clickable((By.ID, 'DtEmissao')))
            time.sleep(2)
            dateEm.click()
            time.sleep(1)
            for _ in range(50):  # Ajuste o número de iterações se necessário
                dateEm.send_keys(Keys.BACKSPACE)
            time.sleep(1)
            dateEm.send_keys(data_formatada)
        
            print('Data inserida com sucesso!')
            time.sleep(5)
        
        data_vencimento = plan['E2'].value

        if isinstance(data_vencimento, datetime):
            datavendimento_forma = data_vencimento.strftime('%d/%m/%Y')
        else:
            print(f"Formato de data inesperado: {data_vencimento}")
            return
        datevem = WebDriverWait(navejgar, 20).until(EC.element_to_be_clickable((By.ID, 'DtVencimento')))
        
        time.sleep(5)
        datevem.click()
        
        time.sleep(1)
        datevem.send_keys(datavendimento_forma)  # Inserir a data
        time.sleep(5)
        
        # Preencher valor bruto
        valor_bruto = WebDriverWait(navejgar, 10).until(EC.element_to_be_clickable((By.ID, 'ValorBruto')))
        valor_bruto.clear()
        valor = float(plan['F2'].value)
        valor_str = f"{valor:,.2f}".replace('.', ',')
        valor_bruto.send_keys(valor_str)

        # Plano de contas
        plano_conta = WebDriverWait(navejgar, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[name$='.PlanoContasText']")))
        plano_conta.clear()
        plano_conta.send_keys(plan['I2'].value + Keys.ENTER)
        ultima_palavra = plan['I2'].value.split('/')[-1].strip()
        planCorrente2 = WebDriverWait(navejgar, 20).until(EC.presence_of_element_located((By.XPATH, f"//td[text()='{ultima_palavra}']")))
        navejgar.execute_script("arguments[0].scrollIntoView(true);", planCorrente2)
        planCorrente2.click()

        # Unidade de custeio
        unidade_custeio = WebDriverWait(navejgar, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[name$='.AreaText']")))
        unidade_custeio.clear()
        unidade_custeio.send_keys(plan['J2'].value + Keys.ENTER)

        # Conta corrente
        conta_corrente = WebDriverWait(navejgar, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[name$='.ContaCorrenteText']")))
        conta_corrente.clear()
        conta_corrente.send_keys(plan['K2'].value + Keys.ENTER)

        # Forma de pagamento
        forma_pagamento = WebDriverWait(navejgar, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[name$='.FormaPagamentoText']")))
        forma_pagamento.clear()
        forma_pagamento.send_keys(plan['L2'].value + Keys.ENTER)

        # Botão salvar
        btn_save = WebDriverWait(navejgar, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[@name='ButtonSave' and @value='0']")))
        btn_save.click()
        time.sleep(5)

        # Verificar pop-ups
        try:
            popup_ok = WebDriverWait(navejgar, 10).until(EC.element_to_be_clickable((By.ID, 'popup_ok')))
            popup_ok.click()
        except:
            pass

    except Exception as e:
        print(f"Erro ao inserir as informações: {e}")
        
def main():
    navejgar = webdriver.Chrome()
    navejgar.get("SEU_SITE_AQUI")

    user = 'Usuario'
    password = 'Login'

    try: 
        print('fazendo o login')
        login(navejgar, user,  password)
        time.sleep(5)
        fechar_pop_up(navejgar)
        for row in plan.iter_rows(min_row=2, values_only=True):
            preencher_campos_areceber(navejgar, plan)
    finally:
        navejgar.quit()

if __name__ == "__main__":
    main() 
